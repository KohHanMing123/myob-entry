import pyautogui as pyag
import time
import sys
from openpyxl import load_workbook
from datetime import datetime


# definitions
def code(beg: int, agent: str):
    extra = 0
    if beg in [40500, 41000] and agent == 'Ex-Clients':
        extra += 30
    t = str(beg + agentno[agent] + extra)
    t = f'{t[0]}-{t[1:]}'
    return t


def countdown(t):
    for i in range(t, 0, -1):
        print(i)
        time.sleep(1)


agentno = {'IM': 1,
           'BB': 2,
           'GE': 3,
           'SM': 4,
           'MD': 5,
           'BH': 6,
           'AL': 7,
           'DS': 8,
           'TP': 9,
           'JB': 10,
           'JW': 11,
           'ML': 13,
           'CD': 12,
           'RS': 14,
           'DD': 15,
           'Ex-Clients': 20}
providers = ['Friends Provident International Limited',
             'FPI (US$)',
             'Utmost International (Quilter)',
             'Utmost Intl (Quilter) USD',
             'Utmost Intl (Quilter) GBP',
             'Utmost International USD',
             'Utmost Worldwide Ltd',
             'Singapore Life Ltd',
             'Monument Intl Life Ass Co Ltd']
beginner = [49100, 49150, 47000, 47100, 40500, 41000, 41016, 44016, 47116, 47516]

# loading data
wb = load_workbook('data.xlsx', data_only=True)
ws = wb.active

# fetching and packing data
data = dict()
collumns = list('ABCDEFGHIJK')
for i in range(len(ws['A'][1:])):
    if ws['A'][i + 1].internal_value is None:
        continue
    d = []
    for letter in collumns:
        d.append(ws[letter][i + 1].internal_value)
    data[i] = tuple(d)


# value: 0=provider 1=invoicedate 2=desc 3=adviser 4=tax 5=v1 6=c1 7=v2 8=c2 9=v3 10=c3
# data validation and confirmation
unknowns = {'Provider': {}, 'Invoice Date': {}, 'Description': {}, 'Adviser': {}, 'Tax': {}, 'Value 1': {},
            'Code 1': {}, 'Value 2': {}, 'Code 2': {}, 'Value 3': {}, 'Code 3': {}}
for rowno, values in data.items():
    rowno += 2
    if values[0] not in providers:
        unknowns['Provider'].setdefault(values[0], []).append(rowno)
    try:
        if len(values[1]) != 8:
            raise Exception
        else:
            datetime(int(values[1][4:]), int(values[1][2:4]), int(values[1][:2])).strftime('%d/%m/%Y')
    except:
        unknowns['Invoice Date'].setdefault(values[1], []).append(rowno)
    if values[2] is None:
        unknowns['Description'].setdefault(values[2], []).append(rowno)
    if values[3] not in agentno:
        unknowns['Adviser'].setdefault(values[3], []).append(rowno)
    if values[4] not in ['ZR', 'SR9', 'SR10']:
        unknowns['Tax'].setdefault(values[4], []).append(rowno)
    if values[5] is None:
        unknowns['Value 1'].setdefault(values[5], []).append(rowno)
    if values[6] is None or len(str(values[6])) != 5:
        unknowns['Code 1'].setdefault(values[6], []).append(rowno)
    if values[7] is not None:
        if values[8] is None or len(str(values[8])) != 5:
            unknowns['Code 2'].setdefault(values[8], []).append(rowno)
    if values[9] is not None:
        if values[10] is None or len(str(values[10])) != 5:
            unknowns['Code 3'].setdefault(values[10], []).append(rowno)
fatalerrors = False
errorno = 0
uerrorno = 0
warnno = 0
uwarnno = 0
for Type, errors in unknowns.items():
    if Type in ['Invoice Date', 'Adviser', 'Value 1', 'Code 1', 'Code 2', 'Code 3']:
        uerrorno += len(errors)
        for error in errors.values():
            errorno += len(error)
    else:
        uwarnno += len(errors)
        for warning in errors.values():
            warnno += len(warning)
input(f'''PYTHON PROGRAM v2: Total number of fatal errors: {errorno}
Number of unique fatal errors: {uerrorno}
Total number of warnings: {warnno}
Number of unique warnings: {uwarnno}

press enter to continue''')
for Type, errors in unknowns.items():
    if len(errors) == 0:
        continue
    if Type in ['Invoice Date', 'Adviser', 'Value 1', 'Code 1', 'Code 2', 'Code 3']:
        print(
            f'fatal error of [{Type}] in the following rows, please amend or remove invalid entry and rerun program '
            f'to continue:')
        fatalerrors = True
        for error, rownos in errors.items():
            error_display = f'"{error}"' if error is not None else 'MISSING VALUE'
            print(f'\t{error_display}:')
            for num in rownos:
                print(f'\t\t\t- {num}')

    print()
if fatalerrors:
    input('finished displaying errors, press enter to close, rerun program after fixing errors')
    sys.exit()
print('\n--------------------------------------------------------\n')
for Type, errors in unknowns.items():
    if len(errors) == 0:
        continue
    if Type not in ['Invoice Date', 'Adviser', 'Value 1', 'Code 1', 'Code 2', 'Code 3']:
        print(f'warnings for [{Type}] in the following rows, please amend before rerunning program or ignore by '
              f'pressing enter for every warning: ')
        for error, rownos in errors.items():
            error_display = f'"{error}"' if error is not None else 'MISSING VALUE'
            print(f'\t{error_display}:', end="")
            for num in rownos:
                print(f'\n\t\t\t- {num}', end="")
            input()
input('\n--------------------------------------------------------\nignore all warnings and proceed?\npress enter to '
      'continue')
countdown(5)

for value in data.values():
    invoicedate = value[1]
    date = datetime(int(invoicedate[4:]), int(invoicedate[2:4]), int(invoicedate[:2])).strftime('%d/%m/%Y')
    bs = f'{value[0]}\n\t\t\t\t\t\t\t\t{invoicedate}\t{date}\t{invoicedate}\t{value[2]}\t{code(value[6], value[3])}\t\t{value[5]}\t\t\t{value[4]}'
    if value[7] is not None and value[8] is not None:
        bs += f'\t\t{value[2]}\t{code(value[8], value[3])}\t\t{value[7]}\t\t\t{value[4]}'
    if value[9] is not None and value[10] is not None:
        bs += f'\t\t{value[2]}\t{code(value[10], value[3])}\t\t{value[9]}\t\t\t{value[4]}'
    pyag.click(161, 83)
    pyag.typewrite(bs)
    if value[4] == 'ZR':
        pyag.click(700, 800)
        pyag.typewrite('ZR')
    pyag.click(180, 795)
    pyag.typewrite(value[2])
    time.sleep(0.5)
    pyag.click(180, 860)
    pyag.typewrite(value[2])
    time.sleep(0.4)
    # Record the entry
    pyag.click(880, 958)
    time.sleep(0.4)
    pyag.click(880, 608)
    #clicks the prompt that comes up when you click the record button
    # pyag.click(700, 962) 
    # the dropdown below the tax column
    # input("clicked record")
    time.sleep(0.5)
input('finished')
